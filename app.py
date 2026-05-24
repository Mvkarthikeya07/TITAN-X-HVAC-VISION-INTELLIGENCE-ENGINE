from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
import os
import json
from datetime import datetime

app = Flask(__name__)

UPLOAD_FOLDER = 'static/uploads'
OUTPUT_FOLDER = 'static/outputs'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Supported file types
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'tif', 'tiff', 'bmp', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_image(filename):
    return filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'tif', 'tiff', 'bmp', 'webp'}

# Lazy-load model (only if weights exist)
_model = None

def get_model():
    global _model
    if _model is None:
        weight_path = "weights/best.pt"
        if os.path.exists(weight_path):
            from models.detector import load_model
            _model = load_model(weight_path)
    return _model


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        return render_template('index.html', error="No file selected.")

    if not allowed_file(file.filename):
        return render_template('index.html', error="Unsupported file type. Please upload a PDF or image.")

    filename = file.filename
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    ext = filename.rsplit('.', 1)[1].lower()
    images = []

    # --- Convert to images ---
    if ext == 'pdf':
        try:
            from models.pdf_processor import pdf_to_images
            images = pdf_to_images(filepath)
        except Exception as e:
            return render_template('index.html', error=f"PDF processing error: {e}")
    else:
        # Already an image — use directly
        images = [filepath]

    total_counts = {}
    output_images = []
    model = get_model()

    # --- Detection ---
    for i, img_path in enumerate(images):
        out_path = os.path.join(OUTPUT_FOLDER, f'output_{i}.jpg')
        output_images.append(f'outputs/output_{i}.jpg')

        if model is not None:
            try:
                from models.detector import detect_objects
                counts = detect_objects(model, img_path, out_path)
            except Exception as e:
                print(f"Detection error on page {i}: {e}")
                counts = {}
        else:
            # Demo mode: simulate detections
            import shutil, random
            shutil.copy(img_path, out_path)
            demo_classes = ['pipe', 'valve', 'duct', 'diffuser', 'pump', 'fan', 'boiler', 'chiller', 'coil', 'filter', 'damper', 'thermostat']
            counts = {cls: random.randint(1, 8) for cls in random.sample(demo_classes, random.randint(4, 8))}

        for k, v in counts.items():
            total_counts[k] = total_counts.get(k, 0) + v

    # --- Cost Calculation ---
    cost_data = calculate_cost(total_counts)

    # --- Excel Export ---
    excel_path = os.path.join(OUTPUT_FOLDER, 'report.xlsx')
    try:
        from utils.excel_export import export_to_excel
        export_to_excel(cost_data, excel_path)
    except Exception as e:
        print(f"Excel export error: {e}")
        export_fallback_excel(cost_data, excel_path)

    # Save session data for dashboard
    session_data = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'filename': filename,
        'counts': total_counts,
        'cost_data': cost_data,
        'images': output_images,
        'pages': len(images)
    }
    with open(os.path.join(OUTPUT_FOLDER, 'session.json'), 'w') as f:
        json.dump(session_data, f)

    return render_template(
        'result.html',
        counts=total_counts,
        cost=cost_data['total_cost'],
        cost_data=cost_data,
        images=output_images,
        excel='outputs/report.xlsx',
        filename=filename,
        pages=len(images),
        timestamp=session_data['timestamp']
    )


@app.route('/download')
def download():
    path = 'static/outputs/report.xlsx'
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name='HVAC_Cost_Report.xlsx')
    return "Report not found. Please run an analysis first.", 404


@app.route('/api/session')
def api_session():
    path = 'static/outputs/session.json'
    if os.path.exists(path):
        with open(path) as f:
            return jsonify(json.load(f))
    return jsonify({})


# ─── COST CALCULATION ───────────────────────────────────────────────────────

UNIT_RATES = {
    'pipe':            15000,
    'valve':           8000,
    'duct':            45000,
    'diffuser':        3500,
    'supply diffuser': 3500,
    'supply_diffuser': 3500,
    'pump':            55000,
    'fan':             28000,
    'boiler':          320000,
    'chiller':         450000,
    'coil':            22000,
    'filter':          6000,
    'damper':          9000,
    'thermostat':      4500,
    # Extended types
    'ahu':             85000,
    'fcu':             22000,
    'vav':             35000,
    'cooling tower':   180000,
    'heat exchanger':  95000,
}

LABOUR_RATE = 0.20   # 20% of materials
OVERHEAD    = 0.10   # 10% overhead

def calculate_cost(counts):
    breakdown = []
    material_total = 0

    for component, qty in counts.items():
        # Normalise key
        key = component.strip().lower().replace('-', '_').replace(' ', '_')
        # Try exact match first, then partial
        rate = UNIT_RATES.get(key) or UNIT_RATES.get(component.lower()) or 10000
        line_total = qty * rate
        material_total += line_total
        breakdown.append({
            'component': component.title(),
            'qty': qty,
            'unit_rate': rate,
            'line_total': line_total
        })

    labour    = int(material_total * LABOUR_RATE)
    overhead  = int(material_total * OVERHEAD)
    total     = material_total + labour + overhead

    return {
        'breakdown': breakdown,
        'material_total': material_total,
        'labour': labour,
        'overhead': overhead,
        'total_cost': total
    }


def export_fallback_excel(cost_data, path):
    """Fallback Excel export using openpyxl directly."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HVAC Cost Report"

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = 'HVAC AI COST ESTIMATION REPORT'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

        # Headers
        headers = ['Component', 'Quantity', 'Unit Rate (₹)', 'Line Total (₹)', 'Notes']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1A4A2A')

        # Data
        for row_i, item in enumerate(cost_data['breakdown'], 5):
            ws.cell(row=row_i, column=1, value=item['component'])
            ws.cell(row=row_i, column=2, value=item['qty'])
            ws.cell(row=row_i, column=3, value=item['unit_rate'])
            ws.cell(row=row_i, column=4, value=item['line_total'])

        last = 5 + len(cost_data['breakdown'])
        ws.cell(row=last,   column=3, value='Materials')
        ws.cell(row=last,   column=4, value=cost_data['material_total']).font = Font(bold=True)
        ws.cell(row=last+1, column=3, value='Labour (20%)')
        ws.cell(row=last+1, column=4, value=cost_data['labour'])
        ws.cell(row=last+2, column=3, value='Overhead (10%)')
        ws.cell(row=last+2, column=4, value=cost_data['overhead'])
        ws.cell(row=last+3, column=3, value='TOTAL')
        ws.cell(row=last+3, column=4, value=cost_data['total_cost']).font = Font(bold=True, size=12)

        for col in [1, 2, 3, 4, 5]:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

        wb.save(path)
    except Exception as e:
        print(f"Fallback Excel also failed: {e}")


if __name__ == '__main__':
    app.run(debug=True, port=5000)