"""
Excel export for HVAC cost report.
Generates a polished .xlsx with formatting, borders, and summary rows.
"""

import os
from datetime import datetime


def export_to_excel(cost_data, output_path):
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HVAC Cost Report"

        # ── Styles ──
        dark_fill   = PatternFill("solid", fgColor="0D2414")
        header_fill = PatternFill("solid", fgColor="1A4A2A")
        alt_fill    = PatternFill("solid", fgColor="0F1E13")
        total_fill  = PatternFill("solid", fgColor="00503A")
        grand_fill  = PatternFill("solid", fgColor="00693A")

        thin = Side(border_style="thin", color="2A6040")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr(text, row, col, bold=True, size=11, color="FFFFFF", fill=header_fill, align="left"):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = Font(bold=bold, size=size, color=color, name="Calibri")
            cell.fill = fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = border
            return cell

        def val(text, row, col, bold=False, color="E0F0E8", fill=None, align="left", number_format=None):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = Font(bold=bold, size=10, color=color, name="Calibri")
            if fill: cell.fill = fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = border
            if number_format: cell.number_format = number_format
            return cell

        # ── Title block ──
        ws.merge_cells("A1:F1")
        title = ws["A1"]
        title.value = "HVAC AI COST ESTIMATION REPORT"
        title.font  = Font(bold=True, size=16, color="00FF6A", name="Calibri")
        title.fill  = dark_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        sub = ws["A2"]
        sub.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  ·  AI-Powered HVAC Component Detection"
        sub.font  = Font(size=9, color="4A7A5A", name="Calibri")
        sub.fill  = dark_fill
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 8  # spacer

        # ── Column headers ──
        headers = ["Component", "Quantity", "Unit Rate (₹)", "Line Total (₹)", "% of Materials", "Status"]
        for c, h in enumerate(headers, 1):
            hdr(h, 4, c, align="center")
        ws.row_dimensions[4].height = 28

        # ── Data rows ──
        breakdown = cost_data.get("breakdown", [])
        mat_total = cost_data.get("material_total", 1) or 1

        for i, item in enumerate(breakdown):
            row = 5 + i
            fill = alt_fill if i % 2 else None
            pct  = round(item["line_total"] / mat_total * 100, 1)

            val(item["component"],  row, 1, fill=fill)
            val(item["qty"],        row, 2, fill=fill, align="center")
            val(item["unit_rate"],  row, 3, fill=fill, align="right", number_format='#,##0')
            val(item["line_total"], row, 4, fill=fill, align="right",  number_format='#,##0', bold=True, color="00FF6A")
            val(f"{pct}%",          row, 5, fill=fill, align="center")
            val("DETECTED",         row, 6, fill=fill, align="center", color="00C952")
            ws.row_dimensions[row].height = 22

        last_data = 5 + len(breakdown)

        # Spacer
        ws.row_dimensions[last_data].height = 6

        # ── Summary rows ──
        def summary_row(label, amount, row, fill=total_fill):
            ws.merge_cells(f"A{row}:C{row}")
            c1 = ws[f"A{row}"]
            c1.value = label
            c1.font  = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            c1.fill  = fill
            c1.alignment = Alignment(horizontal="right", vertical="center")
            c1.border = border
            c2 = ws.cell(row=row, column=4, value=amount)
            c2.font   = Font(bold=True, size=10, color="00FF6A", name="Calibri")
            c2.fill   = fill
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c2.number_format = '₹ #,##0'
            c2.border = border
            for c in [5, 6]:
                cell = ws.cell(row=row, column=c)
                cell.fill = fill
                cell.border = border
            ws.row_dimensions[row].height = 24

        summary_row("Materials Subtotal",   cost_data.get("material_total", 0), last_data + 1)
        summary_row("Labour (20%)",          cost_data.get("labour", 0),         last_data + 2)
        summary_row("Overhead (10%)",        cost_data.get("overhead", 0),       last_data + 3)
        summary_row("★  GRAND TOTAL",        cost_data.get("total_cost", 0),     last_data + 4, fill=grand_fill)
        ws.row_dimensions[last_data + 4].height = 32

        # ── Column widths ──
        widths = [26, 12, 18, 18, 16, 14]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

        # ── Freeze top rows ──
        ws.freeze_panes = "A5"

        # ── Second sheet: raw data ──
        ws2 = wb.create_sheet("Raw Data")
        ws2.append(["Component", "Quantity", "Unit Rate", "Line Total"])
        for item in breakdown:
            ws2.append([item["component"], item["qty"], item["unit_rate"], item["line_total"]])

        wb.save(output_path)
        print(f"[Excel] Report saved: {output_path}")

    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
